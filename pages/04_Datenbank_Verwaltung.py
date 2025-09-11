import streamlit as st
import pandas as pd
import io
from datetime import datetime
from pathlib import Path

# Page Config
st.set_page_config(
    page_title="Datenbank Verwaltung - Ramsperger",
    page_icon="🗄️",
    layout="wide"
)

# ================================================================================================
# BASISKONFIGURATION
# ================================================================================================
BASE_DIR = Path("data")
MASTER_CSV = BASE_DIR / "Ramsperger_Winterreifen_20250826_160010.csv"

# ================================================================================================
# CUSTOM CSS
# ================================================================================================
MAIN_CSS = """
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    
    :root {
        --primary-color: #0ea5e9;
        --primary-dark: #0284c7;
        --secondary-color: #64748b;
        --success-color: #16a34a;
        --warning-color: #f59e0b;
        --error-color: #dc2626;
        --background-light: #f8fafc;
        --background-white: #ffffff;
        --text-primary: #1e293b;
        --text-secondary: #64748b;
        --border-color: #e2e8f0;
        --border-radius: 8px;
        --shadow-sm: 0 1px 2px 0 rgb(0 0 0 / 0.05);
        --shadow-md: 0 4px 6px -1px rgb(0 0 0 / 0.1);
        --shadow-lg: 0 10px 15px -3px rgb(0 0 0 / 0.1);
    }
    
    .main > div {
        padding-top: 1rem;
    }
    
    .main-header {
        background: linear-gradient(135deg, var(--primary-color), var(--primary-dark));
        color: white;
        padding: 2rem;
        border-radius: 12px;
        text-align: center;
        margin-bottom: 2rem;
        box-shadow: var(--shadow-lg);
    }
    
    .main-header h1 {
        margin: 0;
        font-size: 2.5rem;
        font-weight: 700;
        font-family: 'Inter', sans-serif;
    }
    
    .main-header p {
        margin: 0.5rem 0 0 0;
        font-size: 1.1rem;
        opacity: 0.9;
    }
    
    .info-box {
        background: linear-gradient(135deg, #f0fdf4, #dcfce7);
        padding: 1rem;
        border-radius: var(--border-radius);
        border-left: 4px solid var(--success-color);
        margin: 1rem 0;
        box-shadow: var(--shadow-sm);
    }
    
    .warning-box {
        background: linear-gradient(135deg, #fef3c7, #fde68a);
        padding: 1rem;
        border-radius: var(--border-radius);
        border-left: 4px solid var(--warning-color);
        margin: 1rem 0;
        box-shadow: var(--shadow-sm);
    }
    
    .error-box {
        background: linear-gradient(135deg, #fef2f2, #fee2e2);
        padding: 1rem;
        border-radius: var(--border-radius);
        border-left: 4px solid var(--error-color);
        margin: 1rem 0;
        box-shadow: var(--shadow-sm);
    }
    
    [data-testid="metric-container"] {
        background: var(--background-white);
        border: 1px solid var(--border-color);
        padding: 1rem;
        border-radius: var(--border-radius);
        box-shadow: var(--shadow-sm);
    }
    
    .stButton > button {
        border-radius: var(--border-radius);
        border: none;
        font-weight: 500;
        transition: all 0.2s ease;
        font-family: 'Inter', sans-serif;
    }
    
    .stButton > button:hover {
        transform: translateY(-1px);
        box-shadow: var(--shadow-md);
    }
</style>
"""

# CSS anwenden
st.markdown(MAIN_CSS, unsafe_allow_html=True)

# ================================================================================================
# SESSION STATE INITIALISIERUNG
# ================================================================================================
def init_session_state():
    """Initialisiert Session State für Datenbank Verwaltung"""
    if 'db_authenticated' not in st.session_state:
        st.session_state.db_authenticated = False
    if 'db_selected_indices' not in st.session_state:
        st.session_state.db_selected_indices = []

# ================================================================================================
# HELPER FUNCTIONS
# ================================================================================================
def get_efficiency_emoji(rating):
    """Gibt Text für Effizienz-Rating zurück"""
    if pd.isna(rating):
        return ""
    rating = str(rating).strip().upper()[:1]
    return {
        "A": "[A]", "B": "[B]", "C": "[C]", 
        "D": "[D]", "E": "[E]", "F": "[F]", "G": "[G]"
    }.get(rating, "")

def get_stock_display(stock_value):
    """Formatiert Bestandsanzeige mit Text und Emojis"""
    if pd.isna(stock_value) or stock_value == '':
        return "❓ unbekannt"
    
    try:
        stock_num = float(stock_value)
        if stock_num < 0:
            return f"🔴 NACHBESTELLEN ({int(stock_num)})"
        elif stock_num == 0:
            return f"🟡 AUSVERKAUFT ({int(stock_num)})"
        else:
            return f"🟢 VERFÜGBAR ({int(stock_num)})"
    except:
        return "❓ unbekannt"

def clean_dataframe(df):
    """Bereinigt und normalisiert DataFrame"""
    if df.empty:
        return df
    
    # Preis bereinigen
    if "Preis_EUR" in df.columns:
        if df["Preis_EUR"].dtype == object:
            df["Preis_EUR"] = (
                df["Preis_EUR"]
                .astype(str)
                .str.replace(",", ".", regex=False)
                .str.replace("€", "", regex=False)
                .str.strip()
            )
        df["Preis_EUR"] = pd.to_numeric(df["Preis_EUR"], errors="coerce")
    
    # Dimensionen bereinigen
    for c in ["Breite", "Hoehe", "Zoll"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").astype("Int64")
    
    # Bestand als Float (negative Werte erlaubt)
    if "Bestand" in df.columns:
        df["Bestand"] = pd.to_numeric(df["Bestand"], errors="coerce")
    
    # Fehlende Spalten ergänzen
    required_cols = ["Fabrikat", "Profil", "Kraftstoffeffizienz", "Nasshaftung", 
                    "Loadindex", "Speedindex", "Teilenummer", "Geräuschklasse"]
    for col in required_cols:
        if col not in df.columns:
            df[col] = pd.NA
    
    # Leere Zeilen entfernen
    df = df.dropna(subset=["Preis_EUR", "Breite", "Hoehe", "Zoll"], how="any")
    
    if not df.empty:
        df["Breite"] = df["Breite"].astype(int)
        df["Hoehe"] = df["Hoehe"].astype(int)
        df["Zoll"] = df["Zoll"].astype(int)
    
    return df

# ================================================================================================
# DATENBANK FUNKTIONEN
# ================================================================================================
@st.cache_data(show_spinner=False)
def load_master_database():
    """Lädt die Master-Datenbank"""
    if not MASTER_CSV.exists():
        return pd.DataFrame()
    
    try:
        df = pd.read_csv(MASTER_CSV)
        return clean_dataframe(df)
    except Exception as e:
        st.error(f"Fehler beim Laden der Datenbank: {e}")
        return pd.DataFrame()

def save_master_database(df):
    """Speichert die Master-Datenbank"""
    try:
        MASTER_CSV.parent.mkdir(parents=True, exist_ok=True)
        df.to_csv(MASTER_CSV, index=False, encoding='utf-8')
        
        # Cache leeren
        load_master_database.clear()
        
        return True
    except Exception as e:
        st.error(f"Fehler beim Speichern der Datenbank: {e}")
        return False

def update_single_tire(teilenummer, updated_data):
    """Aktualisiert einen einzelnen Reifen"""
    df = load_master_database()
    
    if df.empty:
        return False
    
    # Reifen finden
    mask = df['Teilenummer'] == teilenummer
    if not mask.any():
        return False
    
    # Daten aktualisieren
    for col, value in updated_data.items():
        if col in df.columns:
            df.loc[mask, col] = value
    
    return save_master_database(df)

def remove_tire(teilenummer):
    """Entfernt einen Reifen"""
    df = load_master_database()
    
    if df.empty:
        return False
    
    original_len = len(df)
    df = df[df['Teilenummer'] != teilenummer]
    
    if len(df) < original_len:
        return save_master_database(df)
    
    return False

def mass_update_tires(selected_indices, updates):
    """Führt Massen-Update durch"""
    df = load_master_database()
    
    if df.empty:
        return False
    
    count = 0
    for idx in selected_indices:
        if idx in df.index:
            for update_type, value in updates.items():
                if update_type == 'price_percent' and value != 0:
                    current_price = df.loc[idx, 'Preis_EUR']
                    new_price = current_price * (1 + value / 100)
                    df.loc[idx, 'Preis_EUR'] = new_price
                    count += 1
                elif update_type == 'kraftstoff' and value != 'Nicht ändern':
                    df.loc[idx, 'Kraftstoffeffizienz'] = value
                    count += 1
                elif update_type == 'nasshaftung' and value != 'Nicht ändern':
                    df.loc[idx, 'Nasshaftung'] = value
                    count += 1
                elif update_type == 'geraeusch' and value > 0:
                    df.loc[idx, 'Geräuschklasse'] = value
                    count += 1
                elif update_type == 'bestand' and value != 999:
                    df.loc[idx, 'Bestand'] = value
                    count += 1
    
    if count > 0:
        return save_master_database(df)
    
    return False

# ================================================================================================
# EXPORT FUNKTIONEN
# ================================================================================================
def create_download_csv(df):
    """Erstellt CSV für Download"""
    df_download = df.copy()
    
    # Spalten-Mapping für bessere Lesbarkeit
    column_mapping = {
        'Breite': 'Breite', 'Hoehe': 'Höhe', 'Zoll': 'Zoll', 'Loadindex': 'Tragkraft',
        'Speedindex': 'Geschw.', 'Fabrikat': 'Hersteller', 'Profil': 'Profil',
        'Teilenummer': 'Teilenummer', 'Preis_EUR': 'Preis (EUR)', 'Bestand': 'Bestand',
        'Kraftstoffeffizienz': 'Kraftstoff', 'Nasshaftung': 'Nasshaftung', 
        'Geräuschklasse': 'Lärm (dB)'
    }
    
    df_download = df_download.rename(columns=column_mapping)
    
    # Formatierung
    if 'Preis (EUR)' in df_download.columns:
        df_download['Preis (EUR)'] = df_download['Preis (EUR)'].apply(
            lambda x: f"{x:.2f}".replace('.', ',') if pd.notnull(x) and x != '' else ''
        )
    
    if 'Bestand' in df_download.columns:
        df_download['Bestand'] = df_download['Bestand'].fillna('').apply(
            lambda x: str(int(x)) if pd.notnull(x) and x != '' else ''
        )
    
    if 'Lärm (dB)' in df_download.columns:
        df_download['Lärm (dB)'] = df_download['Lärm (dB)'].fillna('').apply(
            lambda x: str(int(x)) if pd.notnull(x) and x != '' else ''
        )
    
    # Relevante Spalten
    export_columns = ['Breite', 'Höhe', 'Zoll', 'Tragkraft', 'Geschw.', 'Hersteller', 
                     'Profil', 'Teilenummer', 'Preis (EUR)', 'Bestand', 'Kraftstoff', 
                     'Nasshaftung', 'Lärm (dB)']
    available_columns = [col for col in export_columns if col in df_download.columns]
    df_download = df_download[available_columns]
    df_download = df_download.fillna('')
    
    # CSV erstellen
    csv_buffer = io.StringIO()
    df_download.to_csv(csv_buffer, index=False, encoding='utf-8', sep=';', decimal=',')
    return csv_buffer.getvalue()

def create_download_excel(df):
    """Erstellt Excel für Download"""
    df_download = df.copy()
    
    # Formatierung
    if 'Bestand' in df_download.columns:
        df_download['Bestand'] = df_download['Bestand'].fillna('').apply(
            lambda x: int(x) if pd.notnull(x) and x != '' else None
        )
    
    if 'Geräuschklasse' in df_download.columns:
        df_download['Geräuschklasse'] = df_download['Geräuschklasse'].fillna('').apply(
            lambda x: int(x) if pd.notnull(x) and x != '' else None
        )
    
    # Relevante Spalten
    export_columns = ['Breite', 'Hoehe', 'Zoll', 'Loadindex', 'Speedindex', 'Fabrikat', 
                     'Profil', 'Teilenummer', 'Preis_EUR', 'Bestand', 'Kraftstoffeffizienz', 
                     'Nasshaftung', 'Geräuschklasse']
    available_columns = [col for col in export_columns if col in df_download.columns]
    df_download = df_download[available_columns]
    df_download = df_download.replace('', None)
    
    # Excel erstellen
    excel_buffer = io.BytesIO()
    
    try:
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df_download.to_excel(writer, sheet_name='Reifen', index=False)
            worksheet = writer.sheets['Reifen']
            
            # Spaltenbreiten
            column_widths = {1: 10, 2: 10, 3: 8, 4: 12, 5: 10, 6: 15, 7: 25, 8: 18, 
                            9: 12, 10: 8, 11: 16, 12: 12, 13: 14}
            
            for col_idx, width in column_widths.items():
                if col_idx <= len(df_download.columns):
                    col_letter = chr(ord('A') + col_idx - 1)
                    worksheet.column_dimensions[col_letter].width = width
            
            # Header-Formatierung
            try:
                from openpyxl.styles import Font, PatternFill, Alignment
                
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for col_num in range(1, len(df_download.columns) + 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Preis formatieren
                preis_col = None
                for i, col_name in enumerate(df_download.columns, 1):
                    if col_name == 'Preis_EUR':
                        preis_col = chr(ord('A') + i - 1)
                        break
                
                if preis_col:
                    for row in range(2, len(df_download) + 2):
                        cell = worksheet[f'{preis_col}{row}']
                        cell.number_format = '#,##0.00 "€"'
                
                worksheet.auto_filter.ref = f"A1:{chr(ord('A') + len(df_download.columns) - 1)}{len(df_download) + 1}"
                worksheet.freeze_panes = "A2"
                
            except ImportError:
                pass  # Falls openpyxl.styles nicht verfügbar
        
        return excel_buffer.getvalue()
    except Exception as e:
        st.error(f"Excel-Export Fehler: {e}")
        return None

# ================================================================================================
# AUTHENTICATION
# ================================================================================================
def check_authentication():
    """Prüft Authentifizierung für Admin-Bereich"""
    if not st.session_state.db_authenticated:
        st.markdown("""
        <div class="main-header">
            <h1>Datenbank Verwaltung</h1>
            <p>Vollzugriff auf die Reifendatenbank</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("""
        <div class="warning-box">
            <h4>Administratorzugang erforderlich</h4>
            <p>Dieser Bereich ermöglicht vollständige Datenbank-Operationen und ist nur für Administratoren zugänglich.</p>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            password = st.text_input("Administrator-PIN:", type="password", key="db_password")
            
            col_login, col_back = st.columns(2)
            with col_login:
                if st.button("Anmelden", use_container_width=True, type="primary"):
                    if password == "1234":
                        st.session_state.db_authenticated = True
                        st.success("Administratorzugang gewährt!")
                        st.rerun()
                    else:
                        st.error("Falsches Passwort!")
            
            with col_back:
                if st.button("Zurück", use_container_width=True):
                    st.switch_page("app.py")
        
        return False
    return True

# ================================================================================================
# RENDER FUNCTIONS
# ================================================================================================
def render_database_statistics(df):
    """Rendert Datenbank-Statistiken"""
    st.markdown("### 📊 Datenbank Übersicht")
    
    col1, col2, col3, col4, col5 = st.columns(5)
    
    with col1:
        st.metric("Reifen gesamt", len(df))
    
    with col2:
        if not df.empty:
            hersteller_count = df['Fabrikat'].nunique()
            st.metric("Hersteller", hersteller_count)
        else:
            st.metric("Hersteller", 0)
    
    with col3:
        if not df.empty and df['Preis_EUR'].notna().any():
            avg_price = df['Preis_EUR'].mean()
            st.metric("Ø Preis", f"{avg_price:.0f} EUR")
        else:
            st.metric("Ø Preis", "0 EUR")
    
    with col4:
        if 'Bestand' in df.columns and not df.empty:
            with_stock = len(df[df['Bestand'].notna() & (df['Bestand'] > 0)])
            st.metric("Mit Bestand", with_stock)
        else:
            st.metric("Mit Bestand", 0)
    
    with col5:
        if 'Kraftstoffeffizienz' in df.columns and not df.empty:
            with_labels = len(df[df['Kraftstoffeffizienz'].notna() & (df['Kraftstoffeffizienz'] != '')])
            st.metric("Mit EU-Label", with_labels)
        else:
            st.metric("Mit EU-Label", 0)
    
    # Bestandsverteilung
    if 'Bestand' in df.columns and not df.empty:
        render_stock_distribution(df)

def render_stock_distribution(df):
    """Rendert Bestandsverteilung"""
    st.markdown("**📦 Bestandsverteilung:**")
    
    stock_data = df[df['Bestand'].notna()]
    if not stock_data.empty:
        negative = len(stock_data[stock_data['Bestand'] < 0])
        zero = len(stock_data[stock_data['Bestand'] == 0])
        positive = len(stock_data[stock_data['Bestand'] > 0])
        total_stock = stock_data['Bestand'].sum()
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.markdown(f"🔴 **Nachbestellung:** {negative}")
        with col2:
            st.markdown(f"🟡 **Ausverkauft:** {zero}")
        with col3:
            st.markdown(f"🟢 **Verfügbar:** {positive}")
        with col4:
            color = "🔴" if total_stock < 0 else "🟢"
            st.markdown(f"{color} **Gesamtbestand:** {total_stock:.0f}")

def render_tire_list(df):
    """Rendert die Reifen-Liste mit Checkboxes"""
    st.markdown("### 📋 Reifen auswählen")
    
    # Paginierung
    items_per_page = 20
    total_pages = (len(df) + items_per_page - 1) // items_per_page
    
    if total_pages > 1:
        page = st.selectbox(f"Seite (je {items_per_page} Reifen):", range(1, total_pages + 1), key="db_page") - 1
        start_idx = page * items_per_page
        end_idx = min(start_idx + items_per_page, len(df))
        page_data = df.iloc[start_idx:end_idx]
    else:
        page_data = df
    
    # Checkboxes für Auswahl
    for idx, row in page_data.iterrows():
        is_selected = idx in st.session_state.db_selected_indices
        
        col_check, col_info = st.columns([1, 9])
        
        with col_check:
            if st.checkbox("Auswählen", value=is_selected, key=f"db_check_{idx}", label_visibility="hidden"):
                if idx not in st.session_state.db_selected_indices:
                    st.session_state.db_selected_indices.append(idx)
            else:
                if idx in st.session_state.db_selected_indices:
                    st.session_state.db_selected_indices.remove(idx)
        
        with col_info:
            reifengroesse = f"{row['Breite']}/{row['Hoehe']} R{row['Zoll']}"
            
            # Bestandsanzeige
            bestand_info = ""
            if 'Bestand' in row.index:
                bestand_info = f" | {get_stock_display(row['Bestand'])}"
            
            # EU-Label
            eu_info = ""
            if 'Kraftstoffeffizienz' in row.index and pd.notna(row['Kraftstoffeffizienz']) and row['Kraftstoffeffizienz'] != '':
                eu_info += f" {get_efficiency_emoji(row['Kraftstoffeffizienz'])}{row['Kraftstoffeffizienz']}"
            
            st.write(f"**{reifengroesse}** - {row['Fabrikat']} {row['Profil']} - **{row['Preis_EUR']:.2f}EUR**{bestand_info}{eu_info} - {row['Teilenummer']}")

def render_single_tire_editor(df):
    """Rendert Editor für einzelnen Reifen"""
    if len(st.session_state.db_selected_indices) != 1:
        return
    
    st.markdown("---")
    st.markdown("### 🔧 Einzelreifen bearbeiten")
    
    selected_idx = st.session_state.db_selected_indices[0]
    selected_row = df.loc[selected_idx]
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**Reifen Info:**")
        reifengroesse = f"{selected_row['Breite']}/{selected_row['Hoehe']} R{selected_row['Zoll']}"
        st.write(f"**Größe:** {reifengroesse}")
        st.write(f"**Hersteller:** {selected_row['Fabrikat']}")
        st.write(f"**Profil:** {selected_row['Profil']}")
        st.write(f"**Teilenummer:** {selected_row['Teilenummer']}")
        
        # Preis bearbeiten
        new_preis = st.number_input(
            "Preis (EUR):",
            min_value=0.0,
            max_value=2000.0,
            value=float(selected_row['Preis_EUR']),
            step=0.01,
            key="db_edit_preis"
        )
    
    with col2:
        st.markdown("**EU-Labels & Bestand:**")
        
        # Bestand
        current_bestand = selected_row.get('Bestand', 0)
        if pd.isna(current_bestand) or current_bestand == '':
            bestand_value = 0
        else:
            bestand_value = int(current_bestand)
        
        new_bestand = st.number_input(
            "Bestand:",
            min_value=-999,
            max_value=1000,
            value=bestand_value,
            step=1,
            key="db_edit_bestand",
            help="Negative Werte = Nachbestellung nötig"
        )
        
        # Kraftstoffeffizienz
        current_kraftstoff = selected_row.get('Kraftstoffeffizienz', '')
        kraftstoff_options = ['', 'A', 'B', 'C', 'D', 'E', 'F', 'G']
        kraftstoff_index = kraftstoff_options.index(current_kraftstoff) if current_kraftstoff in kraftstoff_options else 0
        
        new_kraftstoff = st.selectbox(
            "Kraftstoffeffizienz:",
            options=kraftstoff_options,
            index=kraftstoff_index,
            key="db_edit_kraftstoff"
        )
        
        # Nasshaftung
        current_nasshaftung = selected_row.get('Nasshaftung', '')
        nasshaftung_index = kraftstoff_options.index(current_nasshaftung) if current_nasshaftung in kraftstoff_options else 0
        
        new_nasshaftung = st.selectbox(
            "Nasshaftung:",
            options=kraftstoff_options,
            index=nasshaftung_index,
            key="db_edit_nasshaftung"
        )
        
        # Geräuschklasse
        current_geraeusch = selected_row.get('Geräuschklasse', 70)
        if pd.isna(current_geraeusch) or current_geraeusch == '':
            geraeusch_value = 70
        else:
            geraeusch_value = int(current_geraeusch)
        
        new_geraeusch = st.number_input(
            "Geräuschklasse (dB):",
            min_value=66,
            max_value=75,
            value=geraeusch_value,
            step=1,
            key="db_edit_geraeusch"
        )
    
    # Buttons
    col_save, col_delete = st.columns(2)
    
    with col_save:
        if st.button("Änderungen speichern", use_container_width=True, type="primary"):
            updated_data = {
                'Preis_EUR': new_preis,
                'Bestand': new_bestand,
                'Kraftstoffeffizienz': new_kraftstoff,
                'Nasshaftung': new_nasshaftung,
                'Geräuschklasse': new_geraeusch if new_geraeusch > 0 else None
            }
            
            if update_single_tire(selected_row['Teilenummer'], updated_data):
                st.success("Reifen erfolgreich aktualisiert!")
                st.rerun()
            else:
                st.error("Fehler beim Aktualisieren!")
    
    with col_delete:
        if st.button("Reifen löschen", use_container_width=True, type="secondary"):
            if remove_tire(selected_row['Teilenummer']):
                st.session_state.db_selected_indices = []
                st.success("Reifen erfolgreich gelöscht!")
                st.rerun()
            else:
                st.error("Fehler beim Löschen!")

def render_mass_editor():
    """Rendert Massen-Bearbeitung"""
    if len(st.session_state.db_selected_indices) <= 1:
        return
    
    st.markdown("---")
    st.markdown(f"### 🔄 Massen-Bearbeitung ({len(st.session_state.db_selected_indices)} Reifen)")
    
    col1, col2 = st.columns(2)
    
    with col1:
        mass_preis_percent = st.number_input(
            "Preise ändern (% Aufschlag/Abschlag):",
            min_value=-50.0,
            max_value=100.0,
            value=0.0,
            step=1.0,
            help="z.B. 10 = +10%, -5 = -5%",
            key="db_mass_preis"
        )
        
        mass_kraftstoff = st.selectbox(
            "Kraftstoffeffizienz für alle setzen:",
            options=['Nicht ändern', '', 'A', 'B', 'C', 'D', 'E', 'F', 'G'],
            key="db_mass_kraftstoff"
        )
    
    with col2:
        mass_nasshaftung = st.selectbox(
            "Nasshaftung für alle setzen:",
            options=['Nicht ändern', '', 'A', 'B', 'C', 'D', 'E', 'F', 'G'],
            key="db_mass_nasshaftung"
        )
        
        mass_geraeusch = st.number_input(
            "Geräuschklasse für alle setzen (0 = nicht ändern):",
            min_value=0,
            max_value=75,
            value=0,
            step=1,
            key="db_mass_geraeusch"
        )
        
        mass_bestand = st.number_input(
            "Bestand für alle setzen (999 = nicht ändern):",
            min_value=-999,
            max_value=1000,
            value=999,
            step=1,
            key="db_mass_bestand",
            help="Negative Werte = Nachbestellung nötig"
        )
    
    # Massen-Update Button
    if st.button("Massen-Update durchführen", use_container_width=True, type="primary"):
        updates = {
            'price_percent': mass_preis_percent,
            'kraftstoff': mass_kraftstoff,
            'nasshaftung': mass_nasshaftung,
            'geraeusch': mass_geraeusch,
            'bestand': mass_bestand
        }
        
        if mass_update_tires(st.session_state.db_selected_indices, updates):
            st.success(f"Massen-Update durchgeführt! {len(st.session_state.db_selected_indices)} Reifen aktualisiert.")
            st.rerun()
        else:
            st.error("Fehler beim Massen-Update!")

def render_export_functions(df, filtered_df):
    """Rendert Export-Funktionen"""
    st.markdown("---")
    st.markdown("### 📁 Export-Funktionen")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if len(filtered_df) > 0:
            csv_data = create_download_csv(filtered_df)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"Ramsperger_Gefilterte_DB_{timestamp}.csv"
            
            st.download_button(
                label="Gefilterte DB als CSV",
                data=csv_data,
                file_name=filename,
                mime="text/csv",
                help="Gefilterte Datenbank als CSV herunterladen"
            )
    
    with col2:
        if len(filtered_df) > 0:
            excel_data = create_download_excel(filtered_df)
            if excel_data:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"Ramsperger_Gefilterte_DB_{timestamp}.xlsx"
                
                st.download_button(
                    label="Gefilterte DB als Excel",
                    data=excel_data,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="Gefilterte Datenbank als Excel herunterladen"
                )
    
    with col3:
        if len(df) > 0:
            csv_data = create_download_csv(df)
            
            st.download_button(
                label="Komplette DB als CSV",
                data=csv_data,
                file_name="Ramsperger_Winterreifen_20250826_160010.csv",
                mime="text/csv",
                help="Komplette Datenbank als CSV herunterladen"
            )

# ================================================================================================
# MAIN FUNCTION
# ================================================================================================
def main():
    init_session_state()
    
    if not check_authentication():
        return
    
    # Header
    st.markdown("""
    <div class="main-header">
        <h1>Datenbank Verwaltung</h1>
        <p>Vollständige Kontrolle über die Reifendatenbank</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Sidebar Navigation
    with st.sidebar:
        st.markdown("---")
        
        if st.button("← Reifen Suche", use_container_width=True):
            st.switch_page("pages/01_Reifen_Suche.py")
        
        if st.button("🛒 Warenkorb", use_container_width=True, type="primary"):
            st.switch_page("pages/02_Warenkorb.py")
        
        if st.button("🔧 Reifen Verwaltung", use_container_width=True):
            st.switch_page("pages/03_Premium_Verwaltung.py")
        
        st.markdown("---")
        st.header("🔍 Datenbank Filter")
        
        # Datenbank laden
        current_df = load_master_database()
        
        if current_df.empty:
            st.error("Keine Datenbank gefunden!")
            return
        
        # Filter-Optionen
        db_fabrikat_opt = ["Alle"] + sorted([x for x in current_df["Fabrikat"].dropna().unique().tolist()])
        db_fabrikat = st.selectbox("Hersteller filtern:", options=db_fabrikat_opt, index=0, key="db_fabrikat")
        
        db_zoll_opt = ["Alle"] + sorted(current_df["Zoll"].unique().tolist())
        db_zoll = st.selectbox("Zoll filtern:", options=db_zoll_opt, index=0, key="db_zoll")
        
        # Teilenummer Suche
        db_search = st.text_input(
            "Teilenummer/Profil suchen:",
            placeholder="z.B. ZTW225 oder WinterContact",
            key="db_search"
        )
        
        st.markdown("---")
        st.header("🛠️ Massen-Aktionen")
        
        # Massen-Auswahl
        if st.button("Alle auswählen", use_container_width=True):
            st.session_state.db_selected_indices = current_df.index.tolist()
            st.rerun()
        
        if st.button("Alle abwählen", use_container_width=True):
            st.session_state.db_selected_indices = []
            st.rerun()
        
        # Gefährliche Aktionen
        st.markdown("**⚠️ Gefährliche Aktionen:**")
        
        if st.button("Ausgewählte löschen", use_container_width=True, type="secondary"):
            if st.session_state.db_selected_indices:
                updated_df = current_df.drop(index=st.session_state.db_selected_indices)
                if save_master_database(updated_df):
                    st.session_state.db_selected_indices = []
                    st.success(f"Reifen gelöscht!")
                    st.rerun()
                else:
                    st.error("Fehler beim Löschen!")
        
        if st.button("🗑️ Komplette DB löschen", use_container_width=True, type="secondary"):
            if save_master_database(pd.DataFrame()):
                st.session_state.db_selected_indices = []
                st.success("Datenbank komplett geleert!")
                st.rerun()
            else:
                st.error("Fehler beim Löschen!")
        
        if st.button("Abmelden", use_container_width=True, type="secondary"):
            st.session_state.db_authenticated = False
            st.rerun()
    
    # Filter anwenden
    filtered_df = current_df.copy()
    
    if db_fabrikat != "Alle":
        filtered_df = filtered_df[filtered_df["Fabrikat"] == db_fabrikat]
    
    if db_zoll != "Alle":
        filtered_df = filtered_df[filtered_df["Zoll"] == int(db_zoll)]
    
    # Bestandsfilter anwenden
    if mit_bestand_filter:
        filtered_df = filtered_df[
            (filtered_df['Bestand'].notna()) & 
            (filtered_df['Bestand'] > 0)
        ]
    
    # Teilenummer-Bulk-Suche
    if db_search and db_search.strip() != "":
        # Komma-getrennte Teilenummern verarbeiten
        search_terms = [term.strip().upper() for term in db_search.split(',') if term.strip()]
        
        if search_terms:
            mask = pd.Series([False] * len(filtered_df))
            
            for search_term in search_terms:
                # Nur in Teilenummer suchen
                term_mask = filtered_df['Teilenummer'].str.upper().str.contains(search_term, na=False, regex=False)
                mask = mask | term_mask
            
            filtered_df = filtered_df[mask]
    
    # Info über Datenbank
    st.markdown("""
    <div class="info-box">
        <h4>📊 Master-Datenbank</h4>
        <p>Du bearbeitest die Haupt-Reifendatenbank. Alle Änderungen sind sofort in der Reifen-Suche sichtbar.</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Statistiken
    render_database_statistics(current_df)
    
    # Hauptbereich
    if len(filtered_df) > 0:
        st.markdown(f"### 📋 Reifen in Datenbank: {len(filtered_df)}")
        
        # Reifen-Liste
        render_tire_list(filtered_df)
        
        # Einzelreifen-Editor
        render_single_tire_editor(current_df)
        
        # Massen-Editor
        render_mass_editor()
        
        # Export-Funktionen
        render_export_functions(current_df, filtered_df)
        
    else:
        st.info("Keine Reifen gefunden. Ändere die Filter oder erstelle Reifen über die Reifen Verwaltung.")

if __name__ == "__main__":
    main()