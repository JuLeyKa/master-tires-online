import streamlit as st
import pandas as pd
import io
from pathlib import Path
from datetime import datetime
import numpy as np

# Page Config
st.set_page_config(
    page_title="Premium Verwaltung - Ramsperger",
    page_icon="⚙️",
    layout="wide"
)

# ================================================================================================
# BASISKONFIGURATION
# ================================================================================================
BASE_DIR = Path("data")  # Für Cloud-Version
MASTER_CSV = BASE_DIR / "Ramsperger_Winterreifen_20250826_160010.csv"
CENTRAL_DATABASE_CSV = BASE_DIR / "ramsperger_central_database.csv"
SERVICES_CONFIG_CSV = BASE_DIR / "ramsperger_services_config.csv"

# ================================================================================================
# CUSTOM CSS
# ================================================================================================
CUSTOM_CSS = """
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    
    :root {
        --primary-color: #0ea5e9;
        --primary-dark: #0284c7;
        --secondary-color: #64748b;
        --success-color: #16a34a;
        --warning-color: #f59e0b;
        --error-color: #dc2626;
        --background-light: #f8fafc;
        --background-white: #ffffff;
        --text-primary: #1e293b;
        --text-secondary: #64748b;
        --border-color: #e2e8f0;
        --border-radius: 8px;
        --shadow-sm: 0 1px 2px 0 rgb(0 0 0 / 0.05);
        --shadow-md: 0 4px 6px -1px rgb(0 0 0 / 0.1);
        --shadow-lg: 0 10px 15px -3px rgb(0 0 0 / 0.1);
    }
    
    .main > div {
        padding-top: 1rem;
    }
    
    .main-header {
        background: linear-gradient(135deg, var(--primary-color), var(--primary-dark));
        color: white;
        padding: 2rem;
        border-radius: 12px;
        text-align: center;
        margin-bottom: 2rem;
        box-shadow: var(--shadow-lg);
    }
    
    .main-header h1 {
        margin: 0;
        font-size: 2.5rem;
        font-weight: 700;
        font-family: 'Inter', sans-serif;
    }
    
    .main-header p {
        margin: 0.5rem 0 0 0;
        font-size: 1.1rem;
        opacity: 0.9;
    }
    
    .workflow-step {
        background: linear-gradient(135deg, #dbeafe, #bfdbfe);
        padding: 1.5rem;
        border-radius: 12px;
        margin: 1rem 0;
        border-left: 4px solid #2563eb;
        font-family: Arial, sans-serif;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
    
    .stats-container {
        background: linear-gradient(135deg, #f0f9ff, #e0f2fe);
        padding: 1rem;
        border-radius: 10px;
        margin: 1rem 0;
        border: 1px solid #0ea5e9;
    }
    
    .filter-info {
        background: linear-gradient(135deg, #fef3c7, #fde68a);
        padding: 1rem;
        border-radius: 8px;
        margin: 1rem 0;
        border-left: 4px solid #f59e0b;
    }
    
    .database-info {
        background: linear-gradient(135deg, #f0fdf4, #dcfce7);
        padding: 1rem;
        border-radius: 8px;
        margin: 1rem 0;
        border-left: 4px solid #16a34a;
    }
    
    .info-box {
        background: linear-gradient(135deg, #f0fdf4, #dcfce7);
        padding: 1rem;
        border-radius: var(--border-radius);
        border-left: 4px solid var(--success-color);
        margin: 1rem 0;
        box-shadow: var(--shadow-sm);
    }
    
    .warning-box {
        background: linear-gradient(135deg, #fef3c7, #fde68a);
        padding: 1rem;
        border-radius: var(--border-radius);
        border-left: 4px solid var(--warning-color);
        margin: 1rem 0;
        box-shadow: var(--shadow-sm);
    }
    
    .error-box {
        background: linear-gradient(135deg, #fef2f2, #fee2e2);
        padding: 1rem;
        border-radius: var(--border-radius);
        border-left: 4px solid var(--error-color);
        margin: 1rem 0;
        box-shadow: var(--shadow-sm);
    }
    
    [data-testid="metric-container"] {
        background: var(--background-white);
        border: 1px solid var(--border-color);
        padding: 1rem;
        border-radius: var(--border-radius);
        box-shadow: var(--shadow-sm);
    }
    
    .stButton > button {
        border-radius: var(--border-radius);
        border: none;
        font-weight: 500;
        transition: all 0.2s ease;
        font-family: 'Inter', sans-serif;
    }
    
    .stButton > button:hover {
        transform: translateY(-1px);
        box-shadow: var(--shadow-md);
    }
</style>
"""

# CSS anwenden
st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

# ================================================================================================
# SESSION STATE INITIALISIERUNG
# ================================================================================================
def init_session_state():
    """Initialisiert den Session State für alle Apps"""
    # Navigation
    if 'current_tab' not in st.session_state:
        st.session_state.current_tab = "Premium Verwaltung"
    if 'authenticated' not in st.session_state:
        st.session_state.authenticated = False
    
    # Tab 2 - Warenkorb/Angebot - ERWEITERT FÜR RADWECHSEL-OPTIONEN
    if 'cart_items' not in st.session_state:
        st.session_state.cart_items = []
    if 'cart_quantities' not in st.session_state:
        st.session_state.cart_quantities = {}
    if 'cart_services' not in st.session_state:
        st.session_state.cart_services = {}
    if 'selected_services' not in st.session_state:
        st.session_state.selected_services = {'montage': False, 'radwechsel': False, 'einlagerung': False}
    
    # Tab 3 - Premium Verwaltung
    if 'df_original' not in st.session_state:
        st.session_state.df_original = None
    if 'df_filtered' not in st.session_state:
        st.session_state.df_filtered = None
    if 'df_working' not in st.session_state:
        st.session_state.df_working = None
    if 'file_uploaded' not in st.session_state:
        st.session_state.file_uploaded = False
    if 'selected_indices' not in st.session_state:
        st.session_state.selected_indices = []
    if 'filter_applied' not in st.session_state:
        st.session_state.filter_applied = False
    if 'selection_confirmed' not in st.session_state:
        st.session_state.selection_confirmed = False
    if 'df_selected' not in st.session_state:
        st.session_state.df_selected = None
    if 'current_tire_index' not in st.session_state:
        st.session_state.current_tire_index = 0
    if 'auto_advance' not in st.session_state:
        st.session_state.auto_advance = True
    if 'services_mode' not in st.session_state:
        st.session_state.services_mode = False
    if 'stock_mode' not in st.session_state:
        st.session_state.stock_mode = False

# ================================================================================================
# SERVICE KONFIGURATION
# ================================================================================================
def load_services_config():
    """Lädt oder erstellt die Service-Konfiguration"""
    if not SERVICES_CONFIG_CSV.exists():
        default_services = pd.DataFrame({
            'service_name': ['montage_bis_17', 'montage_18_19', 'montage_ab_20', 'radwechsel_1_rad', 'radwechsel_2_raeder', 'radwechsel_3_raeder', 'radwechsel_4_raeder', 'nur_einlagerung'],
            'service_label': ['Montage bis 17 Zoll', 'Montage 18-19 Zoll', 'Montage ab 20 Zoll', 'Radwechsel 1 Rad', 'Radwechsel 2 Räder', 'Radwechsel 3 Räder', 'Radwechsel 4 Räder', 'Nur Einlagerung'],
            'price': [25.0, 30.0, 40.0, 9.95, 19.95, 29.95, 39.90, 55.00],
            'unit': ['pro Reifen', 'pro Reifen', 'pro Reifen', 'pauschal', 'pauschal', 'pauschal', 'pauschal', 'pauschal']
        })
        # Erstelle Verzeichnis falls nicht vorhanden
        SERVICES_CONFIG_CSV.parent.mkdir(parents=True, exist_ok=True)
        default_services.to_csv(SERVICES_CONFIG_CSV, index=False, encoding='utf-8')
        return default_services
    else:
        return pd.read_csv(SERVICES_CONFIG_CSV)

def save_services_config(services_df):
    """Speichert die Service-Konfiguration"""
    try:
        SERVICES_CONFIG_CSV.parent.mkdir(parents=True, exist_ok=True)
        services_df.to_csv(SERVICES_CONFIG_CSV, index=False, encoding='utf-8')
        return True
    except Exception as e:
        st.error(f"Fehler beim Speichern der Service-Konfiguration: {e}")
        return False

def get_service_prices():
    """Gibt aktuelle Service-Preise zurück"""
    services_df = load_services_config()
    prices = {}
    for _, row in services_df.iterrows():
        prices[row['service_name']] = row['price']
    return prices

# ================================================================================================
# DATENBANK FUNKTIONEN
# ================================================================================================
def initialize_central_database():
    """Erstellt eine leere zentrale Datenbank falls sie nicht existiert"""
    if not CENTRAL_DATABASE_CSV.exists():
        empty_df = pd.DataFrame(columns=['Breite', 'Hoehe', 'Zoll', 'Loadindex', 'Speedindex', 
                                        'Fabrikat', 'Profil', 'Teilenummer', 'Preis_EUR', 
                                        'Bestand', 'Kraftstoffeffizienz', 'Nasshaftung', 'Geräuschklasse'])
        CENTRAL_DATABASE_CSV.parent.mkdir(parents=True, exist_ok=True)
        empty_df.to_csv(CENTRAL_DATABASE_CSV, index=False, encoding='utf-8')

def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Bereinigt und normalisiert DataFrame - ERWEITERT FÜR NEGATIVE BESTÄNDE"""
    if df.empty:
        return df
    
    if "Preis_EUR" in df.columns:
        if df["Preis_EUR"].dtype == object:
            df["Preis_EUR"] = (
                df["Preis_EUR"]
                .astype(str)
                .str.replace(",", ".", regex=False)
                .str.replace("€", "", regex=False)
                .str.strip()
            )
        df["Preis_EUR"] = pd.to_numeric(df["Preis_EUR"], errors="coerce")

    for c in ["Breite", "Hoehe", "Zoll"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").astype("Int64")

    # Bestand als Float (erlaubt negative Werte!)
    if "Bestand" in df.columns:
        df["Bestand"] = pd.to_numeric(df["Bestand"], errors="coerce")

    for c in ["Fabrikat", "Profil", "Kraftstoffeffizienz", "Nasshaftung", 
              "Loadindex", "Speedindex", "Teilenummer"]:
        if c not in df.columns:
            df[c] = pd.NA

    df = df.dropna(subset=["Preis_EUR", "Breite", "Hoehe", "Zoll"], how="any")
    if not df.empty:
        df["Breite"] = df["Breite"].astype(int)
        df["Hoehe"] = df["Hoehe"].astype(int)
        df["Zoll"] = df["Zoll"].astype(int)

    return df

@st.cache_data(show_spinner=False)
def load_master_csv() -> pd.DataFrame:
    """Lädt die unveränderliche Master-CSV"""
    if not MASTER_CSV.exists():
        return pd.DataFrame()

    df = pd.read_csv(MASTER_CSV)
    return clean_dataframe(df)

@st.cache_data(show_spinner=False)
def load_central_database() -> pd.DataFrame:
    """Lädt die zentrale Arbeits-Datenbank"""
    initialize_central_database()
    
    if not CENTRAL_DATABASE_CSV.exists():
        return pd.DataFrame()

    df = pd.read_csv(CENTRAL_DATABASE_CSV)
    return clean_dataframe(df)

def combine_databases() -> pd.DataFrame:
    """Kombiniert Master-CSV und zentrale DB"""
    master_df = load_master_csv()
    central_df = load_central_database()
    
    if master_df.empty and central_df.empty:
        return pd.DataFrame()
    
    if not master_df.empty and 'Bestand' not in master_df.columns:
        master_df['Bestand'] = pd.NA
    
    if not central_df.empty and 'Bestand' not in central_df.columns:  
        central_df['Bestand'] = pd.NA
    
    if master_df.empty:
        return central_df
    
    if central_df.empty:
        return master_df
    
    if 'Teilenummer' in master_df.columns and 'Teilenummer' in central_df.columns:
        master_teilenummern = set(master_df['Teilenummer'].dropna())
        central_df_filtered = central_df[~central_df['Teilenummer'].isin(master_teilenummern)]
        combined_df = pd.concat([master_df, central_df_filtered], ignore_index=True)
    else:
        combined_df = pd.concat([master_df, central_df], ignore_index=True)
    
    return combined_df

def add_or_update_central_database(new_df):
    """INTELLIGENTE Funktion: Fügt neue Reifen hinzu oder aktualisiert bestehende"""
    try:
        if CENTRAL_DATABASE_CSV.exists():
            existing_df = pd.read_csv(CENTRAL_DATABASE_CSV)
            existing_df = clean_dataframe(existing_df)
        else:
            existing_df = pd.DataFrame()
        
        required_columns = ['Breite', 'Hoehe', 'Zoll', 'Loadindex', 'Speedindex', 'Fabrikat', 
                          'Profil', 'Teilenummer', 'Preis_EUR', 'Bestand', 'Kraftstoffeffizienz', 
                          'Nasshaftung', 'Geräuschklasse']
        
        new_df_clean = new_df.copy()
        for col in required_columns:
            if col not in new_df_clean.columns:
                new_df_clean[col] = ''
        new_df_clean = new_df_clean[required_columns]
        
        if existing_df.empty:
            result_df = new_df_clean
        else:
            if 'Teilenummer' in existing_df.columns and 'Teilenummer' in new_df_clean.columns:
                existing_teilenummern = set(existing_df['Teilenummer'].dropna())
                new_teilenummern = set(new_df_clean['Teilenummer'].dropna())
                
                update_teilenummern = existing_teilenummern.intersection(new_teilenummern)
                add_teilenummern = new_teilenummern - existing_teilenummern
                keep_teilenummern = existing_teilenummern - new_teilenummern
                
                keep_df = existing_df[existing_df['Teilenummer'].isin(keep_teilenummern)]
                add_df = new_df_clean[new_df_clean['Teilenummer'].isin(add_teilenummern)]
                update_df = new_df_clean[new_df_clean['Teilenummer'].isin(update_teilenummern)]
                
                result_df = pd.concat([keep_df, add_df, update_df], ignore_index=True)
            else:
                result_df = pd.concat([existing_df, new_df_clean], ignore_index=True)
        
        CENTRAL_DATABASE_CSV.parent.mkdir(parents=True, exist_ok=True)
        result_df.to_csv(CENTRAL_DATABASE_CSV, index=False, encoding='utf-8')
        return True, len(new_df_clean)
    except Exception as e:
        st.error(f"Fehler beim Aktualisieren der zentralen Datenbank: {e}")
        return False, 0

# ================================================================================================
# PREMIUM EXCEL DATEN LADEN
# ================================================================================================
@st.cache_data(show_spinner=False)
def load_premium_data() -> pd.DataFrame:
    """Lädt die hardcoded Excel für die Premium Verwaltung"""
    excel_path = BASE_DIR / "2025-07-29_ReifenPremium_Winterreifen_2025-26.xlsx"
    
    if not excel_path.exists():
        st.error(f"Excel-Datei nicht gefunden: {excel_path}")
        return pd.DataFrame()
    
    try:
        df = pd.read_excel(excel_path, sheet_name=0)
        
        # Spalten-Namen bereinigen
        df.columns = df.columns.str.replace(r'\r\n', ' ', regex=True).str.strip()
        
        # Dimension zusammenbauen
        df['Dimension'] = (
            df['Breite'].astype(str) + '/' + 
            df['Hoehe'].astype(str) + ' ' + 
            df['R'].astype(str) + df['Zoll'].astype(str) + ' ' + 
            df['Loadindex'].astype(str) + df['Speedindex'].astype(str)
        )
        
        # Runflat-Kennzeichnung
        df['Dimension'] = df.apply(
            lambda row: row['Dimension'] + (' RF' if pd.notna(row['RF']) and row['RF'] != '' else ''), 
            axis=1
        )
        
        # Preis-Spalte finden
        preis_col = None
        for col in df.columns:
            if 'Preis' in col and 'netto' in col:
                preis_col = col
                break
        
        if preis_col:
            df['Preis_EUR'] = pd.to_numeric(df[preis_col], errors='coerce')
        else:
            df['Preis_EUR'] = 0.0
        
        # Spalten umbenennen/erstellen
        required_cols = ['Dimension', 'Fabrikat', 'Profil', 'Teilenummer', 'Preis_EUR', 
                        'Zoll', 'Breite', 'Hoehe', 'RF', 'Kennzeichen']
        for col in required_cols:
            if col not in df.columns:
                df[col] = ''
        
        # Nur relevante Spalten
        df = df[['Dimension', 'Fabrikat', 'Profil', 'Teilenummer', 'Preis_EUR', 'Zoll', 
                'Breite', 'Hoehe', 'RF', 'Kennzeichen', 'Speedindex', 'Loadindex']]
        df = df.fillna('')
        
        return df
    except Exception as e:
        st.error(f"Fehler beim Laden der Excel-Datei: {e}")
        return pd.DataFrame()

def add_new_columns(df):
    """Fügt EU-Label Spalten hinzu"""
    required_original_cols = ['Dimension', 'Fabrikat', 'Profil', 'Teilenummer', 'Preis_EUR']
    for col in required_original_cols:
        if col not in df.columns:
            df[col] = ''
    
    if 'Bestand' not in df.columns:
        df['Bestand'] = pd.Series([None] * len(df), dtype='float64')
    if 'Kraftstoffeffizienz' not in df.columns:
        df['Kraftstoffeffizienz'] = ''
    if 'Nasshaftung' not in df.columns:
        df['Nasshaftung'] = ''
    if 'Geräuschklasse' not in df.columns:
        df['Geräuschklasse'] = pd.Series([None] * len(df), dtype='float64')
    
    return df

# ================================================================================================
# FILTER FUNKTIONEN
# ================================================================================================
def apply_filters(df, hersteller_filter, zoll_filter, preis_range, runflat_filter, 
                 breite_filter, hoehe_filter, teilenummer_search, speed_filter, 
                 stock_filter="alle"):
    """Wendet Sidebar-Filter an - ERWEITERT UM BESTANDSFILTER"""
    filtered_df = df.copy()
    
    if hersteller_filter and len(hersteller_filter) > 0:
        filtered_df = filtered_df[filtered_df['Fabrikat'].isin(hersteller_filter)]
    
    if zoll_filter and len(zoll_filter) > 0:
        filtered_df = filtered_df[filtered_df['Zoll'].isin(zoll_filter)]
    
    filtered_df = filtered_df[
        (filtered_df['Preis_EUR'] >= preis_range[0]) & 
        (filtered_df['Preis_EUR'] <= preis_range[1])
    ]
    
    if runflat_filter == "Nur Runflat":
        filtered_df = filtered_df[filtered_df['RF'] != '']
    elif runflat_filter == "Ohne Runflat":
        filtered_df = filtered_df[filtered_df['RF'] == '']
    
    if breite_filter and len(breite_filter) > 0:
        filtered_df = filtered_df[filtered_df['Breite'].isin(breite_filter)]
    
    if hoehe_filter and len(hoehe_filter) > 0:
        filtered_df = filtered_df[filtered_df['Hoehe'].isin(hoehe_filter)]
    
    if teilenummer_search and teilenummer_search.strip() != "":
        search_terms = [term.strip().upper() for term in teilenummer_search.split(',') if term.strip()]
        
        if search_terms:
            mask = pd.Series([False] * len(filtered_df))
            
            for search_term in search_terms:
                term_mask = (
                    filtered_df['Teilenummer'].str.upper().str.contains(search_term, na=False, regex=False) |
                    filtered_df['Fabrikat'].str.upper().str.contains(search_term, na=False, regex=False) |
                    filtered_df['Profil'].str.upper().str.contains(search_term, na=False, regex=False)
                )
                mask = mask | term_mask
            
            filtered_df = filtered_df[mask]
    
    if speed_filter and len(speed_filter) > 0:
        filtered_df = filtered_df[filtered_df['Speedindex'].isin(speed_filter)]
    
    # NEU: BESTANDSFILTER
    if stock_filter == "negative_only":
        # Nur Reifen mit negativem Bestand (Nachbestellung nötig)
        filtered_df = filtered_df[
            (filtered_df['Bestand'].notna()) & 
            (filtered_df['Bestand'] < 0)
        ]
    elif stock_filter == "zero_or_negative":
        # Reifen mit Bestand <= 0
        filtered_df = filtered_df[
            (filtered_df['Bestand'].notna()) & 
            (filtered_df['Bestand'] <= 0)
        ]
    elif stock_filter == "positive_only":
        # Nur Reifen mit positivem Bestand
        filtered_df = filtered_df[
            (filtered_df['Bestand'].notna()) & 
            (filtered_df['Bestand'] > 0)
        ]
    elif stock_filter == "no_stock_info":
        # Reifen ohne Bestandsinfo
        filtered_df = filtered_df[filtered_df['Bestand'].isna()]
    
    return filtered_df

# ================================================================================================
# BESTANDSMANAGEMENT
# ================================================================================================
def get_stock_statistics(df):
    """Berechnet Bestandsstatistiken"""
    stats = {}
    
    if 'Bestand' not in df.columns:
        return {'total': len(df), 'with_stock': 0, 'negative': 0, 'zero': 0, 'positive': 0, 'no_info': len(df)}
    
    # Basis-Statistiken
    stats['total'] = len(df)
    stats['with_stock'] = len(df[df['Bestand'].notna()])
    stats['no_info'] = len(df[df['Bestand'].isna()])
    
    # Bestandsverteilung
    stock_data = df[df['Bestand'].notna()]
    stats['negative'] = len(stock_data[stock_data['Bestand'] < 0])
    stats['zero'] = len(stock_data[stock_data['Bestand'] == 0])
    stats['positive'] = len(stock_data[stock_data['Bestand'] > 0])
    
    # Gesamtbestand berechnen
    stats['total_stock'] = stock_data['Bestand'].sum()
    
    return stats

def create_reorder_list_export(df):
    """Erstellt Export für Nachbestellliste"""
    if df.empty:
        return "Keine Nachbestellungen erforderlich."
    
    content = []
    content.append("RAMSPERGER REIFEN - NACHBESTELLLISTE")
    content.append("=" * 60)
    content.append(f"Erstellt: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")
    content.append("")
    
    # Negative Bestände nach Hersteller gruppieren
    grouped = df.groupby('Fabrikat')
    
    for fabrikat, group in grouped:
        content.append(f"HERSTELLER: {fabrikat}")
        content.append("-" * 40)
        
        for _, row in group.iterrows():
            reifengroesse = f"{row['Breite']}/{row['Hoehe']} R{row['Zoll']}"
            bestand = int(row['Bestand']) if pd.notna(row['Bestand']) else 0
            nachbedarf = abs(bestand) if bestand < 0 else 0
            
            content.append(f"• {reifengroesse} - {row['Profil']}")
            content.append(f"  Teilenummer: {row['Teilenummer']}")
            content.append(f"  Aktueller Bestand: {bestand}")
            content.append(f"  Nachbedarf: {nachbedarf} Stück")
            content.append(f"  Einzelpreis: {row['Preis_EUR']:.2f}€")
            content.append(f"  Gesamtwert: {nachbedarf * row['Preis_EUR']:.2f}€")
            content.append("")
        
        content.append("")
    
    # Zusammenfassung
    total_nachbedarf = 0
    total_wert = 0.0
    
    for _, row in df.iterrows():
        bestand = int(row['Bestand']) if pd.notna(row['Bestand']) else 0
        if bestand < 0:
            nachbedarf = abs(bestand)
            total_nachbedarf += nachbedarf
            total_wert += nachbedarf * row['Preis_EUR']
    
    content.append("=" * 60)
    content.append("ZUSAMMENFASSUNG")
    content.append("-" * 30)
    content.append(f"Reifen-Typen mit Nachbedarf: {len(df)}")
    content.append(f"Gesamt Nachbedarf: {total_nachbedarf} Stück")
    content.append(f"Geschätzter Bestellwert: {total_wert:.2f}€")
    content.append("=" * 60)
    
    return "\n".join(content)

# ================================================================================================
# EXPORT FUNKTIONEN
# ================================================================================================
def create_download_excel(df):
    """Erstellt Excel für Download"""
    df_download = df.copy()
    
    if 'Bestand' in df_download.columns:
        df_download['Bestand'] = df_download['Bestand'].fillna('').apply(
            lambda x: int(x) if pd.notnull(x) and x != '' else None
        )
    
    if 'Geräuschklasse' in df_download.columns:
        df_download['Geräuschklasse'] = df_download['Geräuschklasse'].fillna('').apply(
            lambda x: int(x) if pd.notnull(x) and x != '' else None
        )
    
    required_columns = ['Breite', 'Hoehe', 'Zoll', 'Loadindex', 'Speedindex', 'Fabrikat', 
                       'Profil', 'Teilenummer']
    for col in required_columns:
        if col not in df_download.columns:
            df_download[col] = ''
    
    export_columns = ['Breite', 'Hoehe', 'Zoll', 'Loadindex', 'Speedindex', 'Fabrikat', 
                     'Profil', 'Teilenummer', 'Preis_EUR', 'Bestand', 'Kraftstoffeffizienz', 
                     'Nasshaftung', 'Geräuschklasse']
    available_columns = [col for col in export_columns if col in df_download.columns]
    df_download = df_download[available_columns]
    df_download = df_download.replace('', None)
    
    excel_buffer = io.BytesIO()
    
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df_download.to_excel(writer, sheet_name='Reifen', index=False)
        worksheet = writer.sheets['Reifen']
        
        # Spaltenbreiten
        column_widths = {1: 10, 2: 10, 3: 8, 4: 12, 5: 10, 6: 15, 7: 25, 8: 18, 
                        9: 12, 10: 8, 11: 16, 12: 12, 13: 14}
        
        for col_idx, width in column_widths.items():
            if col_idx <= len(df_download.columns):
                col_letter = chr(ord('A') + col_idx - 1)
                worksheet.column_dimensions[col_letter].width = width
        
        # Formatierung
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col_num in range(1, len(df_download.columns) + 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Preis formatieren
            preis_col = None
            for i, col_name in enumerate(df_download.columns, 1):
                if col_name == 'Preis_EUR':
                    preis_col = chr(ord('A') + i - 1)
                    break
            
            if preis_col:
                for row in range(2, len(df_download) + 2):
                    cell = worksheet[f'{preis_col}{row}']
                    cell.number_format = '#,##0.00 "€"'
            
            worksheet.auto_filter.ref = f"A1:{chr(ord('A') + len(df_download.columns) - 1)}{len(df_download) + 1}"
            worksheet.freeze_panes = "A2"
        except ImportError:
            pass  # Falls openpyxl.styles nicht verfügbar
    
    return excel_buffer.getvalue()

def create_download_csv(df):
    """Erstellt CSV für Download"""
    df_download = df.copy()
    
    column_mapping = {
        'Breite': 'Breite', 'Hoehe': 'Höhe', 'Zoll': 'Zoll', 'Loadindex': 'Tragkraft',
        'Speedindex': 'Geschw.', 'Fabrikat': 'Hersteller', 'Profil': 'Profil',
        'Teilenummer': 'Teilenummer', 'Preis_EUR': 'Preis (EUR)', 'Bestand': 'Bestand',
        'Kraftstoffeffizienz': 'Kraftstoff', 'Nasshaftung': 'Nasshaftung', 
        'Geräuschklasse': 'Lärm (dB)'
    }
    
    df_download = df_download.rename(columns=column_mapping)
    
    if 'Preis (EUR)' in df_download.columns:
        df_download['Preis (EUR)'] = df_download['Preis (EUR)'].apply(
            lambda x: f"{x:.2f}".replace('.', ',') if pd.notnull(x) and x != '' else ''
        )
    
    if 'Bestand' in df_download.columns:
        df_download['Bestand'] = df_download['Bestand'].fillna('').apply(
            lambda x: str(int(x)) if pd.notnull(x) and x != '' else ''
        )
    
    if 'Lärm (dB)' in df_download.columns:
        df_download['Lärm (dB)'] = df_download['Lärm (dB)'].fillna('').apply(
            lambda x: str(int(x)) if pd.notnull(x) and x != '' else ''
        )
    
    required_columns = ['Breite', 'Höhe', 'Zoll', 'Tragkraft', 'Geschw.', 'Hersteller', 
                       'Profil', 'Teilenummer']
    for col in required_columns:
        if col not in df_download.columns:
            df_download[col] = ''
    
    export_columns = ['Breite', 'Höhe', 'Zoll', 'Tragkraft', 'Geschw.', 'Hersteller', 
                     'Profil', 'Teilenummer', 'Preis (EUR)', 'Bestand', 'Kraftstoff', 
                     'Nasshaftung', 'Lärm (dB)']
    available_columns = [col for col in export_columns if col in df_download.columns]
    df_download = df_download[available_columns]
    df_download = df_download.fillna('')
    
    csv_buffer = io.StringIO()
    df_download.to_csv(csv_buffer, index=False, encoding='utf-8', sep=';', decimal=',')
    return csv_buffer.getvalue()

# NEU: Vollständige Datenbank Export Funktion
def create_complete_database_export():
    """Erstellt vollständige Datenbank-CSV (Master + Central) für GitHub Update"""
    try:
        # Beide Datenbanken laden und intelligent kombinieren
        combined_df = combine_databases()
        
        if combined_df.empty:
            return None
        
        # Spalten für GitHub-kompatible CSV vorbereiten
        required_columns = ['Breite', 'Hoehe', 'Zoll', 'Loadindex', 'Speedindex', 'Fabrikat', 
                           'Profil', 'Teilenummer', 'Preis_EUR', 'Bestand', 'Kraftstoffeffizienz', 
                           'Nasshaftung', 'Geräuschklasse']
        
        # Fehlende Spalten ergänzen
        for col in required_columns:
            if col not in combined_df.columns:
                combined_df[col] = ''
        
        # Nur relevante Spalten exportieren
        export_df = combined_df[required_columns].copy()
        
        # Leere Werte durch NaN ersetzen für saubere CSV
        export_df = export_df.replace('', pd.NA)
        
        # CSV erstellen
        csv_buffer = io.StringIO()
        export_df.to_csv(csv_buffer, index=False, encoding='utf-8', na_rep='')
        
        return csv_buffer.getvalue()
        
    except Exception as e:
        st.error(f"Fehler beim Erstellen des vollständigen Datenbank-Exports: {e}")
        return None

# ================================================================================================
# AUTHENTICATION
# ================================================================================================
def check_authentication():
    """Prüft Authentifizierung für Admin-Bereich"""
    if not st.session_state.authenticated:
        st.markdown("""
        <div class="main-header">
            <h1>Premium Verwaltung</h1>
            <p>Passwort-geschützter Adminbereich</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("""
        <div class="warning-box">
            <h4>Authentifizierung erforderlich</h4>
            <p>Dieser Bereich ist nur für autorisierte Benutzer zugänglich.</p>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            password = st.text_input("PIN eingeben:", type="password", key="premium_password")
            
            col_login, col_back = st.columns(2)
            with col_login:
                if st.button("Anmelden", use_container_width=True, type="primary"):
                    if password == "1234":  # Standard-Passwort
                        st.session_state.authenticated = True
                        st.success("Zugang gewährt!")
                        st.rerun()
                    else:
                        st.error("Falsches Passwort!")
            
            with col_back:
                if st.button("Zurück", use_container_width=True):
                    st.switch_page("pages/01_Reifen_Suche.py")
        
        return False
    return True

# ================================================================================================
# SERVICE MANAGEMENT
# ================================================================================================
def render_services_management():
    """Service-Preise Verwaltung - KOMPLETT NEU GESCHRIEBEN"""
    st.markdown("#### ⚙️ Service-Preise verwalten")
    st.markdown("Hier können die Preise für Montage, Radwechsel und Einlagerung angepasst werden.")
    
    # Services laden
    services_df = load_services_config()
    
    # Aktuelle Service-Preise anzeigen
    st.markdown("**Aktuelle Service-Preise:**")
    
    # Service-Preise in einem Dictionary organisieren
    current_prices = {}
    for _, row in services_df.iterrows():
        current_prices[row['service_name']] = float(row['price'])
    
    # Eingabefelder für alle Services
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**Montage-Preise:**")
        
        montage_17 = st.number_input(
            "Montage bis 17 Zoll (€ pro Reifen):",
            min_value=0.0,
            max_value=100.0,
            value=current_prices.get('montage_bis_17', 25.0),
            step=0.10,
            key="service_montage_17"
        )
        
        montage_18 = st.number_input(
            "Montage 18-19 Zoll (€ pro Reifen):",
            min_value=0.0,
            max_value=100.0,
            value=current_prices.get('montage_18_19', 30.0),
            step=0.10,
            key="service_montage_18"
        )
        
        montage_20 = st.number_input(
            "Montage ab 20 Zoll (€ pro Reifen):",
            min_value=0.0,
            max_value=100.0,
            value=current_prices.get('montage_ab_20', 40.0),
            step=0.10,
            key="service_montage_20"
        )
    
    with col2:
        st.markdown("**Radwechsel & Einlagerung:**")
        
        radwechsel_1 = st.number_input(
            "Radwechsel 1 Rad (€):",
            min_value=0.0,
            max_value=50.0,
            value=current_prices.get('radwechsel_1_rad', 9.95),
            step=0.05,
            key="service_radwechsel_1"
        )
        
        radwechsel_2 = st.number_input(
            "Radwechsel 2 Räder (€):",
            min_value=0.0,
            max_value=50.0,
            value=current_prices.get('radwechsel_2_raeder', 19.95),
            step=0.05,
            key="service_radwechsel_2"
        )
        
        radwechsel_3 = st.number_input(
            "Radwechsel 3 Räder (€):",
            min_value=0.0,
            max_value=50.0,
            value=current_prices.get('radwechsel_3_raeder', 29.95),
            step=0.05,
            key="service_radwechsel_3"
        )
        
        radwechsel_4 = st.number_input(
            "Radwechsel 4 Räder (€):",
            min_value=0.0,
            max_value=100.0,
            value=current_prices.get('radwechsel_4_raeder', 39.90),
            step=0.10,
            key="service_radwechsel_4"
        )
        
        einlagerung = st.number_input(
            "Nur Einlagerung (€ pauschal):",
            min_value=0.0,
            max_value=200.0,
            value=current_prices.get('nur_einlagerung', 55.00),
            step=0.10,
            key="service_einlagerung"
        )
    
    # Speichern Button
    if st.button("💾 Preise speichern", use_container_width=True, type="primary"):
        # DataFrame aktualisieren
        services_df.loc[services_df['service_name'] == 'montage_bis_17', 'price'] = montage_17
        services_df.loc[services_df['service_name'] == 'montage_18_19', 'price'] = montage_18
        services_df.loc[services_df['service_name'] == 'montage_ab_20', 'price'] = montage_20
        services_df.loc[services_df['service_name'] == 'radwechsel_1_rad', 'price'] = radwechsel_1
        services_df.loc[services_df['service_name'] == 'radwechsel_2_raeder', 'price'] = radwechsel_2
        services_df.loc[services_df['service_name'] == 'radwechsel_3_raeder', 'price'] = radwechsel_3
        services_df.loc[services_df['service_name'] == 'radwechsel_4_raeder', 'price'] = radwechsel_4
        services_df.loc[services_df['service_name'] == 'nur_einlagerung', 'price'] = einlagerung
        
        if save_services_config(services_df):
            st.success("Service-Preise erfolgreich aktualisiert!")
            st.rerun()
        else:
            st.error("Fehler beim Speichern der Service-Preise!")
    
    # Zurück zu Reifen-Verwaltung
    if st.button("🔧 Zur Reifen-Verwaltung", use_container_width=True):
        st.session_state.services_mode = False
        st.rerun()

# ================================================================================================
# STOCK MANAGEMENT
# ================================================================================================
def render_stock_management():
    """NEU: Bestandsmanagement & Nachbestellungen"""
    st.markdown("#### 📦 Bestandsmanagement & Nachbestellungen")
    st.markdown("Überblick über Lagerbestände und automatische Nachbestelllisten.")
    
    # Kombinierte Daten aus allen Datenbanken laden
    all_data_df = combine_databases()
    
    if all_data_df.empty:
        st.warning("Keine Daten für Bestandsanalysis verfügbar.")
        return
    
    # Bestandsstatistiken berechnen
    stats = get_stock_statistics(all_data_df)
    
    st.markdown("**📊 Bestandsübersicht:**")
    
    col1, col2, col3, col4, col5 = st.columns(5)
    with col1:
        st.metric("Reifen gesamt", stats['total'])
    with col2:
        st.metric("Mit Bestandsinfo", stats['with_stock'])
    with col3:
        if stats['negative'] > 0:
            st.metric("🔴 Nachbestellung", stats['negative'])
        else:
            st.metric("✅ Keine Nachbestellung", stats['negative'])
    with col4:
        st.metric("🟡 Bestand = 0", stats['zero'])
    with col5:
        st.metric("🟢 Verfügbar", stats['positive'])
    
    if stats['total_stock'] < 0:
        st.error(f"⚠️ Gesamtbestand: {stats['total_stock']:.0f} (Negative Bilanz!)")
    else:
        st.success(f"✅ Gesamtbestand: {stats['total_stock']:.0f}")
    
    # Nachbestellliste generieren
    if stats['negative'] > 0:
        st.markdown("---")
        st.markdown("#### ⚠️ Reifen mit Nachbedarf")
        
        negative_stock_df = all_data_df[
            (all_data_df['Bestand'].notna()) & 
            (all_data_df['Bestand'] < 0)
        ].copy()
        
        if not negative_stock_df.empty:
            # Nachbestellliste anzeigen
            st.markdown(f"**{len(negative_stock_df)} Reifen-Typen benötigen Nachbestellung:**")
            
            # Nach Nachbedarf sortieren (größter Rückstand zuerst)
            negative_stock_df = negative_stock_df.sort_values('Bestand')
            
            for idx, row in negative_stock_df.head(10).iterrows():  # Top 10 anzeigen
                reifengroesse = f"{row['Breite']}/{row['Hoehe']} R{row['Zoll']}"
                bestand = int(row['Bestand'])
                nachbedarf = abs(bestand)
                
                col_info, col_details = st.columns([3, 1])
                with col_info:
                    st.markdown(f"🔴 **{reifengroesse}** - {row['Fabrikat']} {row['Profil']}")
                    st.markdown(f"Teilenummer: {row['Teilenummer']} | Einzelpreis: {row['Preis_EUR']:.2f}€")
                with col_details:
                    st.metric("Rückstand", f"{nachbedarf} Stück")
                    st.metric("Bestellwert", f"{nachbedarf * row['Preis_EUR']:.2f}€")
            
            if len(negative_stock_df) > 10:
                st.info(f"... und {len(negative_stock_df) - 10} weitere Reifen mit Nachbedarf")
            
            # Export-Funktionen für Nachbestellliste
            st.markdown("---")
            st.markdown("**📥 Nachbestellliste exportieren:**")
            
            col_export1, col_export2, col_export3 = st.columns(3)
            
            with col_export1:
                # Text-Export der Nachbestellliste
                if st.button("📄 Nachbestellliste (TXT)", use_container_width=True):
                    export_content = create_reorder_list_export(negative_stock_df)
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"Nachbestellliste_Ramsperger_{timestamp}.txt"
                    
                    st.download_button(
                        label="📥 Download Nachbestellliste",
                        data=export_content,
                        file_name=filename,
                        mime="text/plain"
                    )
            
            with col_export2:
                # Excel-Export der Nachbestellliste
                try:
                    excel_data = create_download_excel(negative_stock_df)
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"Nachbestellliste_Ramsperger_{timestamp}.xlsx"
                    
                    st.download_button(
                        label="📊 Nachbestellliste (Excel)",
                        data=excel_data,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                except Exception as e:
                    st.error(f"Excel-Export Fehler: {e}")
            
            with col_export3:
                # CSV-Export der Nachbestellliste
                csv_data = create_download_csv(negative_stock_df)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"Nachbestellliste_Ramsperger_{timestamp}.csv"
                
                st.download_button(
                    label="📋 Nachbestellliste (CSV)",
                    data=csv_data,
                    file_name=filename,
                    mime="text/csv"
                )
    else:
        st.success("✅ Alle Reifen sind ausreichend auf Lager!")
    
    # Zurück zu Reifen-Verwaltung
    if st.button("🔧 Zur Reifen-Verwaltung", use_container_width=True):
        st.session_state.services_mode = False
        st.session_state.stock_mode = False
        st.rerun()

# ================================================================================================
# MAIN PREMIUM CONTENT
# ================================================================================================
def render_premium_content():
    """Hauptinhalt der Premium Verwaltung - ERWEITERT UM BESTANDSFILTER"""
    
    # Sicherstellen, dass Excel-Daten geladen sind
    if st.session_state.df_original is None or st.session_state.df_original.empty:
        with st.spinner('Lade Excel-Daten...'):
            df_premium = load_premium_data()
            if not df_premium.empty:
                st.session_state.df_original = df_premium.copy()
                st.session_state.file_uploaded = True
                st.session_state.filter_applied = False
                st.session_state.selection_confirmed = False
                st.rerun()
            else:
                st.error("Fehler beim Laden der Excel-Datei. Bitte prüfe ob die Datei '2025-07-29_ReifenPremium_Winterreifen_2025-26.xlsx' im data/ Ordner existiert.")
                return
    
    # Sidebar Status und Filter
    with st.sidebar:
        # Workflow-Status
        st.header("Workflow-Status")
        if not st.session_state.filter_applied:
            st.info("Schritt 1: Filter setzen")
            if st.session_state.df_original is not None:
                st.success(f"{len(st.session_state.df_original)} Reifen geladen")
        elif not st.session_state.selection_confirmed:
            st.info("Schritt 2: Reifen auswählen")
            if st.session_state.df_filtered is not None:
                st.success(f"{len(st.session_state.df_filtered)} Reifen gefiltert")
        else:
            st.success("Schritt 3: Bearbeitung")
            if st.session_state.df_working is not None:
                st.success(f"{len(st.session_state.df_working)} Reifen ausgewählt")
        
        # Filter nur anzeigen wenn passend
        if st.session_state.df_original is not None and not st.session_state.selection_confirmed:
            st.header("Intelligente Filter")
            
            df_orig = st.session_state.df_original
            
            # Filter-Komponenten
            alle_hersteller = sorted(df_orig['Fabrikat'].unique())
            hersteller_filter = st.multiselect(
                "Hersteller wählen:",
                options=alle_hersteller,
                default=[],
                key="hersteller_filter"
            )
            
            alle_zolle = sorted(df_orig['Zoll'].unique())
            zoll_filter = st.multiselect(
                "Zoll-Größen:",
                options=alle_zolle,
                default=[],
                key="zoll_filter"
            )
            
            # Preisfilter - VEREINFACHT OHNE SLIDER
            st.markdown("**Preisfilter:**")
            min_preis_input = st.number_input(
                "Mindestpreis (€):",
                min_value=0.0,
                max_value=10000.0,
                value=0.0,
                step=10.0,
                key="min_preis_input"
            )
            max_preis_input = st.number_input(
                "Höchstpreis (€):",
                min_value=0.0,
                max_value=10000.0,
                value=1000.0,
                step=10.0,
                key="max_preis_input"
            )
            preis_range = (min_preis_input, max_preis_input)
            
            runflat_filter = st.selectbox(
                "Runflat-Reifen:",
                options=["Alle", "Nur Runflat", "Ohne Runflat"],
                key="runflat_filter"
            )
            
            alle_breiten = sorted(df_orig['Breite'].unique())
            breite_filter = st.multiselect(
                "Reifenbreite:",
                options=alle_breiten,
                default=[],
                key="breite_filter"
            )
            
            alle_hoehen = sorted(df_orig['Hoehe'].unique())
            hoehe_filter = st.multiselect(
                "Reifenhöhe:",
                options=alle_hoehen,
                default=[],
                key="hoehe_filter"
            )
            
            teilenummer_search = st.text_input(
                "Teilenummer suchen:",
                placeholder="z.B. ZTW225, Continental, WinterContact (mehrere durch Komma trennen)",
                help="Suche in Teilenummer, Hersteller oder Profil. Mehrere Begriffe durch Komma trennen.",
                key="teilenummer_search"
            )
            
            alle_speed = sorted(df_orig['Speedindex'].unique())
            speed_filter = st.multiselect(
                "Geschwindigkeitsindex:",
                options=alle_speed,
                default=[],
                key="speed_filter"
            )
            
            # NEU: BESTANDSFILTER
            st.markdown("---")
            st.markdown("**📦 Bestandsfilter:**")
            stock_filter_options = [
                ("alle", "Alle Reifen"),
                ("negative_only", "🔴 Nur negative Bestände (Nachbedarf)"),
                ("zero_or_negative", "🟡 Bestand ≤ 0"),
                ("positive_only", "🟢 Nur positive Bestände"),
                ("no_stock_info", "❓ Ohne Bestandsinfo")
            ]
            
            stock_filter = st.selectbox(
                "Bestand filtern:",
                options=[opt[0] for opt in stock_filter_options],
                format_func=lambda x: next(opt[1] for opt in stock_filter_options if opt[0] == x),
                key="stock_filter"
            )
            
            if st.button("Filter anwenden", use_container_width=True, type="primary"):
                filtered_df = apply_filters(
                    df_orig, hersteller_filter, zoll_filter, preis_range, 
                    runflat_filter, breite_filter, hoehe_filter, teilenummer_search, 
                    speed_filter, stock_filter
                )
                st.session_state.df_filtered = filtered_df
                st.session_state.filter_applied = True
                st.session_state.selected_indices = []
                st.rerun()
            
            if st.button("Filter zurücksetzen"):
                st.session_state.filter_applied = False
                st.session_state.df_filtered = None
                st.session_state.selected_indices = []
                st.rerun()
        
        # Filter-Info bei Schritt 2/3
        if st.session_state.filter_applied and st.session_state.df_filtered is not None:
            st.markdown("**Aktive Filter:**")
            st.write(f"{len(st.session_state.df_filtered)} von {len(st.session_state.df_original)} Reifen")
    
    # STUFE 1: Filter-Anwendung
    if not st.session_state.filter_applied:
        st.markdown("### Schritt 1: Intelligente Filter setzen")
        st.markdown(f"Aus {len(st.session_state.df_original)} Reifen die gewünschten herausfiltern")
        
        col1, col2, col3, col4 = st.columns(4)
        df_orig = st.session_state.df_original
        
        with col1:
            st.metric("Gesamt Reifen", len(df_orig))
        with col2:
            st.metric("Hersteller", df_orig['Fabrikat'].nunique())
        with col3:
            st.metric("Zoll-Größen", df_orig['Zoll'].nunique())
        with col4:
            avg_preis = df_orig[df_orig['Preis_EUR'] > 0]['Preis_EUR'].mean()
            st.metric("Durchschnittspreis", f"{avg_preis:.0f} Euro")
        
        st.info("Setze deine Filter in der Sidebar und klicke 'Filter anwenden'")
    
    # STUFE 2: Reifen-Auswahl aus gefilterten Daten
    elif st.session_state.filter_applied and not st.session_state.selection_confirmed:
        st.markdown("### Schritt 2: Gefilterte Reifen auswählen")
        st.markdown(f"Wähle aus den {len(st.session_state.df_filtered)} gefilterten Reifen deine gewünschten aus")
        
        df_filtered = st.session_state.df_filtered
        
        if len(df_filtered) == 0:
            st.warning("Keine Reifen gefunden! Bitte Filter anpassen.")
            if st.button("Zurück zu Schritt 1"):
                st.session_state.filter_applied = False
                st.rerun()
        else:
            # Schnell-Auswahl Buttons
            st.markdown("**Schnell-Auswahl:**")
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                if st.button("Alle auswählen"):
                    st.session_state.selected_indices = df_filtered.index.tolist()
                    st.rerun()
            
            with col2:
                if st.button("Alle abwählen"):
                    st.session_state.selected_indices = []
                    st.rerun()
            
            with col3:
                if st.button("Top 20 günstigste"):
                    cheapest = df_filtered.nsmallest(20, 'Preis_EUR')
                    st.session_state.selected_indices = cheapest.index.tolist()
                    st.rerun()
            
            with col4:
                if st.button("Zufällig 10"):
                    random_selection = df_filtered.sample(min(10, len(df_filtered)))
                    st.session_state.selected_indices = random_selection.index.tolist()
                    st.rerun()
            
            # Auswahl-Statistiken
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Gefiltert", len(df_filtered))
            with col2:
                st.metric("Ausgewählt", len(st.session_state.selected_indices))
            with col3:
                if len(st.session_state.selected_indices) > 0:
                    selected_df = df_filtered.loc[st.session_state.selected_indices]
                    avg_preis = selected_df['Preis_EUR'].mean()
                    st.metric("Durchschnittspreis Auswahl", f"{avg_preis:.0f} Euro")
            
            # Gefilterte Reifen-Liste mit Checkboxes
            st.markdown("**Reifen einzeln auswählen:**")
            
            # Paginierung
            items_per_page = 50
            total_pages = (len(df_filtered) + items_per_page - 1) // items_per_page
            
            if total_pages > 1:
                page = st.selectbox(f"Seite (je {items_per_page} Reifen):", range(1, total_pages + 1)) - 1
                start_idx = page * items_per_page
                end_idx = min(start_idx + items_per_page, len(df_filtered))
                page_data = df_filtered.iloc[start_idx:end_idx]
            else:
                page_data = df_filtered
            
            # Checkboxes für aktuelle Seite
            for idx, row in page_data.iterrows():
                is_selected = idx in st.session_state.selected_indices
                
                col_check, col_info = st.columns([1, 9])
                
                with col_check:
                    if st.checkbox("Auswählen", value=is_selected, key=f"check_{idx}", label_visibility="hidden"):
                        if idx not in st.session_state.selected_indices:
                            st.session_state.selected_indices.append(idx)
                    else:
                        if idx in st.session_state.selected_indices:
                            st.session_state.selected_indices.remove(idx)
                
                with col_info:
                    runflat_info = " **RF**" if row['RF'] != '' else ""
                    
                    # ERWEITERT: Bestandsanzeige mit Symbolen
                    bestand_info = ""
                    if 'Bestand' in row.index and pd.notna(row['Bestand']):
                        bestand = float(row['Bestand'])
                        if bestand < 0:
                            bestand_info = f" 🔴({int(bestand)})"
                        elif bestand == 0:
                            bestand_info = f" 🟡({int(bestand)})"
                        elif bestand > 0:
                            bestand_info = f" 🟢({int(bestand)})"
                    
                    st.write(f"**{row['Dimension']}**{runflat_info} - {row['Fabrikat']} {row['Profil']} - **{row['Preis_EUR']:.2f}€**{bestand_info} - {row['Teilenummer']}")
            
            # Auswahl bestätigen
            if len(st.session_state.selected_indices) > 0:
                if st.button("Auswahl bestätigen & weiter zu Schritt 3", use_container_width=True, type="primary"):
                    st.session_state.df_selected = df_filtered.loc[st.session_state.selected_indices].copy()
                    st.session_state.df_selected = add_new_columns(st.session_state.df_selected)
                    st.session_state.df_working = st.session_state.df_selected.copy()
                    st.session_state.selection_confirmed = True
                    st.rerun()
            else:
                st.warning("Bitte mindestens einen Reifen auswählen!")
            
            if st.button("Zurück zu Filter-Einstellungen"):
                st.session_state.filter_applied = False
                st.rerun()
    
    # STUFE 3: Bearbeitung der ausgewählten Reifen
    elif st.session_state.selection_confirmed and st.session_state.df_working is not None:
        st.markdown("### Schritt 3: EU-Labels hinzufügen & Preise anpassen")
        st.markdown(f"Bearbeite die {len(st.session_state.df_working)} ausgewählten Reifen")
        
        # Anzeige-Tabelle
        st.markdown("#### Ausgewählte Reifen")
        
        display_df = st.session_state.df_working.copy()
        
        display_columns = ['Breite', 'Hoehe', 'Zoll', 'Loadindex', 'Speedindex', 'Fabrikat', 
                          'Profil', 'Teilenummer', 'Preis_EUR', 'Bestand', 'Kraftstoffeffizienz', 
                          'Nasshaftung', 'Geräuschklasse']
        available_display_columns = [col for col in display_columns if col in display_df.columns]
        display_df = display_df[available_display_columns]
        
        if 'Bestand' in display_df.columns:
            display_df['Bestand'] = display_df['Bestand'].fillna('').apply(
                lambda x: str(int(x)) if pd.notnull(x) and x != '' else '-'
            )
        if 'Kraftstoffeffizienz' in display_df.columns:
            display_df['Kraftstoffeffizienz'] = display_df['Kraftstoffeffizienz'].fillna('').apply(
                lambda x: x if x != '' else '-'
            )
        if 'Nasshaftung' in display_df.columns:
            display_df['Nasshaftung'] = display_df['Nasshaftung'].fillna('').apply(
                lambda x: x if x != '' else '-'
            )
        if 'Geräuschklasse' in display_df.columns:
            display_df['Geräuschklasse'] = display_df['Geräuschklasse'].fillna('').apply(
                lambda x: f"{int(x)} dB" if pd.notnull(x) and x != '' else '-'
            )
        
        st.dataframe(display_df, use_container_width=True, hide_index=True)
        
        # Einzelnen Reifen bearbeiten
        st.markdown("#### Einzelnen Reifen bearbeiten")
        
        if len(st.session_state.df_working) > 0:
            max_index = len(st.session_state.df_working) - 1
            if st.session_state.current_tire_index > max_index:
                st.session_state.current_tire_index = max_index
            if st.session_state.current_tire_index < 0:
                st.session_state.current_tire_index = 0
            
            # Navigation-Controls
            col_nav1, col_nav2, col_nav3, col_nav4, col_nav5 = st.columns([2, 1, 1, 1, 2])
            
            with col_nav1:
                st.info(f"Reifen {st.session_state.current_tire_index + 1} von {len(st.session_state.df_working)}")
            
            with col_nav2:
                if st.button("< Vorheriger", disabled=(st.session_state.current_tire_index == 0)):
                    st.session_state.current_tire_index -= 1
                    st.rerun()
            
            with col_nav3:
                if st.button("Nächster >", disabled=(st.session_state.current_tire_index >= max_index)):
                    st.session_state.current_tire_index += 1
                    st.rerun()
            
            with col_nav5:
                st.session_state.auto_advance = st.checkbox("Auto-Advance", value=st.session_state.auto_advance, 
                                                          help="Automatisch zum nächsten Reifen nach dem Speichern")
            
            # Reifen-Liste für Dropdown
            reifen_options = []
            df_working_list = list(st.session_state.df_working.iterrows())
            
            for i, (idx, row) in enumerate(df_working_list):
                option_text = f"{i+1}: {row['Dimension']} - {row['Fabrikat']} {row['Profil']} ({row['Preis_EUR']:.2f}€)"
                reifen_options.append((option_text, i))
            
            # Dropdown
            selected_dropdown_index = st.selectbox(
                "Oder Reifen aus Liste auswählen:",
                options=range(len(reifen_options)),
                index=st.session_state.current_tire_index,
                format_func=lambda x: reifen_options[x][0],
                key="reifen_select"
            )
            
            if selected_dropdown_index != st.session_state.current_tire_index:
                st.session_state.current_tire_index = selected_dropdown_index
                st.rerun()
            
            # Aktuellen Reifen holen
            current_position = st.session_state.current_tire_index
            selected_idx, selected_row = df_working_list[current_position]
            
            # Bearbeitungsbereich
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("**Reifen Info:**")
                st.write(f"**Dimension:** {selected_row['Dimension']}")
                st.write(f"**Hersteller:** {selected_row['Fabrikat']}")
                st.write(f"**Profil:** {selected_row['Profil']}")
                st.write(f"**Teilenummer:** {selected_row['Teilenummer']}")
                
                new_preis = st.number_input(
                    "Preis:",
                    min_value=0.0,
                    max_value=2000.0,
                    value=float(selected_row['Preis_EUR']),
                    step=0.01,
                    key=f"preis_{selected_idx}"
                )
            
            with col2:
                st.markdown("**EU-Labels & Bestand:**")
                
                current_bestand = selected_row.get('Bestand', None)
                if pd.isna(current_bestand) or current_bestand == '':
                    bestand_value = 0
                else:
                    bestand_value = int(current_bestand)
                    
                # ERWEITERT: Bestand kann jetzt negativ werden!
                new_bestand = st.number_input(
                    "Bestand:",
                    min_value=-999,  # Negative Bestände erlaubt!
                    max_value=1000,
                    value=bestand_value,
                    step=1,
                    key=f"bestand_{selected_idx}",
                    help="Negative Werte = Nachbestellung nötig"
                )
                
                current_kraftstoff = selected_row.get('Kraftstoffeffizienz', '')
                kraftstoff_options = ['', 'A', 'B', 'C', 'D', 'E', 'F', 'G']
                kraftstoff_index = kraftstoff_options.index(current_kraftstoff) if current_kraftstoff in kraftstoff_options else 0
                
                new_kraftstoff = st.selectbox(
                    "Kraftstoffeffizienz:",
                    options=kraftstoff_options,
                    index=kraftstoff_index,
                    key=f"kraftstoff_{selected_idx}"
                )
                
                current_nasshaftung = selected_row.get('Nasshaftung', '')
                nasshaftung_index = kraftstoff_options.index(current_nasshaftung) if current_nasshaftung in kraftstoff_options else 0
                
                new_nasshaftung = st.selectbox(
                    "Nasshaftung:",
                    options=kraftstoff_options,
                    index=nasshaftung_index,
                    key=f"nasshaftung_{selected_idx}"
                )
                
                current_geraeusch = selected_row.get('Geräuschklasse', None)
                if pd.isna(current_geraeusch) or current_geraeusch == '':
                    geraeusch_value = 70
                else:
                    geraeusch_value = int(current_geraeusch)
                    
                new_geraeusch = st.number_input(
                    "Geräuschklasse (dB):",
                    min_value=66,
                    max_value=75,
                    value=geraeusch_value,
                    step=1,
                    key=f"geraeusch_{selected_idx}"
                )
            
            # Buttons
            col_save, col_remove = st.columns(2)
            
            with col_save:
                if st.button("Änderungen speichern", use_container_width=True, type="primary"):
                    st.session_state.df_working.loc[selected_idx, 'Preis_EUR'] = new_preis
                    st.session_state.df_working.loc[selected_idx, 'Bestand'] = new_bestand  # Kann jetzt negativ sein!
                    st.session_state.df_working.loc[selected_idx, 'Kraftstoffeffizienz'] = new_kraftstoff
                    st.session_state.df_working.loc[selected_idx, 'Nasshaftung'] = new_nasshaftung
                    st.session_state.df_working.loc[selected_idx, 'Geräuschklasse'] = new_geraeusch if new_geraeusch > 0 else None
                    
                    success, count = add_or_update_central_database(st.session_state.df_working)
                    if success:
                        # Cache leeren
                        load_master_csv.clear()
                        load_central_database.clear()
                        
                        if st.session_state.auto_advance and st.session_state.current_tire_index < len(st.session_state.df_working) - 1:
                            st.session_state.current_tire_index += 1
                            st.success(f"Änderungen gespeichert & {count} Reifen zur Datenbank hinzugefügt/aktualisiert! Automatisch zu Reifen {st.session_state.current_tire_index + 1} gewechselt.")
                        else:
                            st.success(f"Änderungen gespeichert & {count} Reifen zur Datenbank hinzugefügt/aktualisiert!")
                    else:
                        st.error("Fehler beim Aktualisieren der Datenbank!")
                    
                    st.rerun()
            
            with col_remove:
                if st.button("Reifen entfernen", use_container_width=True, type="secondary"):
                    st.session_state.df_working = st.session_state.df_working.drop(index=[selected_idx])
                    
                    if selected_idx in st.session_state.selected_indices:
                        st.session_state.selected_indices.remove(selected_idx)
                    
                    if st.session_state.current_tire_index >= len(st.session_state.df_working):
                        st.session_state.current_tire_index = max(0, len(st.session_state.df_working) - 1)
                    
                    st.success(f"Reifen aus Bearbeitung entfernt! Noch {len(st.session_state.df_working)} Reifen in der Liste.")
                    st.rerun()
        else:
            st.warning("Keine Reifen mehr vorhanden!")
        
        # Action Buttons
        col_btn1, col_btn2, col_btn3, col_btn4, col_btn5 = st.columns(5)
        
        with col_btn1:
            if st.button("Zurück zur Auswahl", help="Neue Reifen auswählen"):
                st.session_state.selection_confirmed = False
                st.session_state.current_tire_index = 0
                st.rerun()
        
        with col_btn2:
            if st.button("Daten zurücksetzen", help="Alle Änderungen verwerfen"):
                st.session_state.df_working = st.session_state.df_selected.copy()
                st.session_state.current_tire_index = 0
                st.rerun()
        
        with col_btn3:
            if len(st.session_state.df_working) > 0:
                csv_data = create_download_csv(st.session_state.df_working)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"Ramsperger_Winterreifen_Export_{timestamp}.csv"
                
                st.download_button(
                    label="CSV Export",
                    data=csv_data,
                    file_name=filename,
                    mime="text/csv",
                    help="Fertige Reifen-Tabelle als CSV herunterladen"
                )
            else:
                st.info("Keine Reifen zum Download vorhanden")
        
        with col_btn4:
            if len(st.session_state.df_working) > 0:
                try:
                    excel_data = create_download_excel(st.session_state.df_working)
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"Ramsperger_Winterreifen_Export_{timestamp}.xlsx"
                    
                    st.download_button(
                        label="Excel Export",
                        data=excel_data,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        help="Fertige Reifen-Tabelle als Excel herunterladen"
                    )
                except Exception as e:
                    st.error(f"Excel-Export Fehler: {e}")
            else:
                st.info("Keine Reifen zum Download vorhanden")
        
        with col_btn5:
            if st.button("Alle zur Datenbank hinzufügen", help="Alle bearbeiteten Reifen zur zentralen Datenbank hinzufügen"):
                if len(st.session_state.df_working) > 0:
                    success, count = add_or_update_central_database(st.session_state.df_working)
                    if success:
                        load_master_csv.clear()
                        load_central_database.clear()
                        st.success(f"Alle {count} Reifen erfolgreich zur Datenbank hinzugefügt/aktualisiert!")
                    else:
                        st.error("Fehler beim Speichern zur Datenbank!")
                else:
                    st.warning("Keine Reifen zum Speichern vorhanden!")
        
        # NEU: Vollständige Datenbank Export - UNAUFFÄLLIG AM ENDE
        st.markdown("---")
        st.markdown("#### 🔄 Vollständige Datenbank für GitHub Update")
        
        complete_db_data = create_complete_database_export()
        if complete_db_data:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"Ramsperger_Winterreifen_VOLLSTAENDIG_{timestamp}.csv"
            
            col_info, col_download = st.columns([2, 1])
            with col_info:
                st.info("Lädt die komplette Datenbank (Master-CSV + alle neuen Reifen) für das GitHub Update herunter.")
            with col_download:
                st.download_button(
                    label="📥 Vollständige DB herunterladen",
                    data=complete_db_data,
                    file_name=filename,
                    mime="text/csv",
                    help="Vollständige Datenbank für GitHub Update",
                    use_container_width=True
                )
        else:
            st.warning("Keine Daten für DB-Export verfügbar")
        
        st.markdown("---")
        st.info("🔄 **Intelligentes System:** Neue Reifen werden hinzugefügt, bestehende Reifen (gleiche Teilenummer) werden aktualisiert. Keine Daten gehen verloren!")
        st.info("📦 **Bestandsmanagement:** Negative Bestände zeigen Nachbedarf an. Nutze das Bestandsmanagement für Nachbestelllisten!")

# ================================================================================================
# MAIN TAB RENDER FUNCTION
# ================================================================================================
def render_premium_tab():
    """Hauptfunktion für Premium Verwaltung Tab"""
    st.markdown("### Premium Reifen Verwaltung")
    
    # Sidebar Zurück-Button und Status
    with st.sidebar:
        st.markdown("---")
        if st.button("← Zurück zur Reifen Suche", use_container_width=True):
            st.switch_page("pages/01_Reifen_Suche.py")
        
        if st.button("🛒 Zum Warenkorb", use_container_width=True, type="primary"):
            st.switch_page("pages/02_Warenkorb.py")
        
        if st.button("🗄️ Datenbank Verwaltung", use_container_width=True, type="secondary"):
            st.switch_page("pages/04_Datenbank_Verwaltung.py")
        
        # Modus-Auswahl - ERWEITERT UM BESTANDSMANAGEMENT
        st.markdown("---")
        st.header("Verwaltungsmodus")
        
        modus_options = ["Reifen Verwaltung", "Service-Preise", "Bestandsmanagement"]
        
        if st.session_state.services_mode:
            current_modus = "Service-Preise"
        elif getattr(st.session_state, 'stock_mode', False):
            current_modus = "Bestandsmanagement"
        else:
            current_modus = "Reifen Verwaltung"
        
        new_modus = st.selectbox(
            "Modus wählen:",
            options=modus_options,
            index=modus_options.index(current_modus),
            key="premium_modus_select"
        )
        
        if new_modus != current_modus:
            st.session_state.services_mode = (new_modus == "Service-Preise")
            st.session_state.stock_mode = (new_modus == "Bestandsmanagement")
            st.rerun()
    
    # Modus-spezifischer Content
    if st.session_state.services_mode:
        render_services_management()
    elif getattr(st.session_state, 'stock_mode', False):
        render_stock_management()
    else:
        render_premium_content()

# ================================================================================================
# MAIN FUNCTION
# ================================================================================================
def main():
    init_session_state()
    
    if not check_authentication():
        return
    
    # Header
    st.markdown("""
    <div class="main-header">
        <h1>Premium Verwaltung</h1>
        <p>Erweiterte Reifen- und Systemverwaltung</p>
    </div>
    """, unsafe_allow_html=True)
    
    render_premium_tab()

if __name__ == "__main__":
    main()